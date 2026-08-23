"""
GL Journal Processor — core logic.

All functions here are deterministic and side-effect-free with respect to
the filesystem — they operate on bytes/strings in, structured data or
strings out. This is intentional: Vercel Functions are stateless and have
no persistent disk, so nothing here reads or writes a file path. The one
LLM-driven judgment step (proposing Account for unmatched rows) lives
entirely outside this module, in whatever agent/service calls it.
"""

from __future__ import annotations

import csv
import io
import re
from typing import BinaryIO, Literal, TypedDict

import openpyxl


# --------------------------------------------------------------------------
# Constants
# --------------------------------------------------------------------------

CSV_COLUMNS = [
    "Narration", "Date", "Description", "Account", "TaxRate", "Amount",
    "TrackingName1", "TrackingOption1", "TrackingName2", "TrackingOption2",
]

TAX_RATE = "BAS Excluded"
GL_DATA_START_ROW = 6  # fallback first data row, if a header row can't be located
PRIOR_JOURNAL_REQUIRED_COLUMNS = ("Description", "Account")

LedgerFormat = Literal["xero", "myob"]
LEDGER_FORMATS: tuple[LedgerFormat, ...] = ("xero", "myob")

# Description separator convention differs by source system:
#   Xero:  "412 - Accounting"
#   MYOB:  "1-1000 Cheque account"
_DESCRIPTION_SEPARATOR: dict[LedgerFormat, str] = {
    "xero": " - ",
    "myob": " ",
}


# --------------------------------------------------------------------------
# Types
# --------------------------------------------------------------------------

class GLRow(TypedDict):
    description: str
    amount: str


class MatchedRow(GLRow):
    account_code: str


# --------------------------------------------------------------------------
# Errors
# --------------------------------------------------------------------------

class ProcessingError(Exception):
    """Raised for any expected, user-facing failure (bad input, wrong
    format, etc). The API layer catches this and returns a 4xx with the
    message intact, as opposed to an unexpected exception (500, generic)."""


# --------------------------------------------------------------------------
# Pure logic (no I/O — unit-testable in isolation)
# --------------------------------------------------------------------------

def format_amount(debit: float, credit: float) -> str:
    """Apply the debit/credit formatting convention: net = debit - credit;
    negative net is wrapped in parentheses, positive net is shown as-is."""
    net = (debit or 0) - (credit or 0)
    return f"({abs(net):,.2f})" if net < 0 else f"{net:,.2f}"


def match_rows(
    gl_rows: list[GLRow], prior_lookup: dict[str, str]
) -> tuple[list[MatchedRow], list[GLRow]]:
    """Split GL rows into matched (exact Description match against the
    prior journal) and unmatched (needs judgment upstream)."""
    matched: list[MatchedRow] = []
    unmatched: list[GLRow] = []
    for row in gl_rows:
        account_code = prior_lookup.get(row["description"])
        if account_code is not None:
            matched.append({**row, "account_code": account_code})
        else:
            unmatched.append(row)
    return matched, unmatched


def build_csv_rows(rows: list[GLRow] | list[MatchedRow], year: int) -> list[dict[str, str]]:
    """Turn GL rows into full CSV template rows, filling in the
    fixed/derived fields (Narration, Date, TaxRate, blank tracking fields).

    Works for both matched rows (which already have an `account_code`) and
    unmatched rows (which don't yet) — Account is left blank ("") for
    any row missing that key, ready for a downstream step to fill in.
    """
    narration = f"{year}-01 Load Net Activity"
    date = f"30 Jun {year}"
    return [
        {
            "Narration": narration,
            "Date": date,
            "Description": row["description"],
            "Account": row.get("account_code", ""),
            "TaxRate": TAX_RATE,
            "Amount": row["amount"],
            "TrackingName1": "",
            "TrackingOption1": "",
            "TrackingName2": "",
            "TrackingOption2": "",
        }
        for row in rows
    ]


def rows_to_csv_string(rows: list[dict[str, str]]) -> str:
    """Serialize already-shaped CSV rows (keyed by CSV_COLUMNS) to a CSV
    string in memory. This is the only 'write' step in a serverless
    deployment — the result is returned directly in the HTTP response
    body, never written to disk."""
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=CSV_COLUMNS)
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue()


# --------------------------------------------------------------------------
# Parsing (still no filesystem access — operates on bytes/str already in memory)
# --------------------------------------------------------------------------

def parse_ledger(
    file: BinaryIO | bytes,
    filename: str | None = None,
    ledger_format: LedgerFormat | None = None,
) -> tuple[list[GLRow], int | None, LedgerFormat]:
    """Read a GL summary .xlsx (Xero) or GL report .xlsx (MYOB) — as an
    in-memory byte stream, e.g. an uploaded file's contents — into
    normalized rows plus the financial year's start year (parsed from a
    period line near the top of the sheet).

    Which layout to expect is picked automatically: the header row is
    located first ('Account Code' => Xero, 'Category name' => MYOB), and
    if that's inconclusive the filename is used as a fallback signal
    (Xero's default export is named 'General Ledger Summary', MYOB's
    'General Ledger Report'). Pass `ledger_format` to override detection
    entirely.

    `file` can be raw bytes or any binary file-like object (e.g. FastAPI's
    UploadFile.file) — never a filesystem path, since this must work with
    no persistent disk.

    Returns (rows, year, detected_ledger_format).
    """
    stream = io.BytesIO(file) if isinstance(file, (bytes, bytearray)) else file

    try:
        workbook = openpyxl.load_workbook(stream, data_only=True)
    except Exception as exc:  # noqa: BLE001 — re-raised as a clear, typed error
        raise ProcessingError(f"Could not read GL summary file: {exc}") from exc

    sheet = workbook.active
    year = _extract_year(sheet)

    header = _find_header_row(sheet)
    detected_format = ledger_format or (header[1] if header else None) \
        or _detect_format_from_filename(filename)
    if detected_format not in LEDGER_FORMATS:
        raise ProcessingError(
            "Could not determine whether this is a Xero or MYOB GL export. "
            "Expected a header row containing 'Account Code' (Xero) or "
            "'Category name' (MYOB), or a filename indicating the source "
            "system. Pass ledger_format explicitly to override detection."
        )

    data_start_row = header[0] + 1 if header else GL_DATA_START_ROW
    row_parser = _ROW_PARSERS[detected_format]

    results: list[GLRow] = []
    for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
        parsed = row_parser(row)
        if parsed is not None:
            results.append(parsed)

    if not results:
        raise ProcessingError(
            "No account rows found in the uploaded GL summary. Confirm "
            "this is a GL summary export in the expected layout."
        )

    return results, year, detected_format


def _parse_xero_row(row: tuple) -> GLRow | None:
    """Xero GL summary row layout: Account | Account Code | Debit | Credit
    | Net Movement | Account Type."""
    if len(row) < 5:
        return None
    account, code, debit, credit, *_rest = row
    if not account or not code:
        return None
    if str(account).strip().lower() == "total":
        return None
    return {
        "description": f"{code}{_DESCRIPTION_SEPARATOR['xero']}{account}",
        "amount": format_amount(debit or 0, credit or 0),
    }


def _parse_myob_row(row: tuple) -> GLRow | None:
    """MYOB GL report row layout: Code | Category name | Open | Debit |
    Credit | Net activity | Balance | Tax amount. MYOB inserts a blank
    row between every account row, and description is code+name joined
    by a plain space rather than Xero's ' - '."""
    if len(row) < 5:
        return None
    code, category, _open, debit, credit, *_rest = row
    if not code or not category:
        return None
    if str(code).strip().lower() in ("total", "grand total"):
        return None
    return {
        "description": f"{code}{_DESCRIPTION_SEPARATOR['myob']}{category}",
        "amount": format_amount(debit or 0, credit or 0),
    }


_ROW_PARSERS = {
    "xero": _parse_xero_row,
    "myob": _parse_myob_row,
}


def _find_header_row(sheet) -> tuple[int, LedgerFormat] | None:
    """Scan the top of the sheet for the column-header row, returning its
    1-indexed row number plus which format it matches. Returns None if
    neither format's header text is found (caller falls back to filename)."""
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        cells = [str(c).strip().lower() for c in row if isinstance(c, str)]
        if "account code" in cells:
            return idx, "xero"
        if "category name" in cells:
            return idx, "myob"
    return None


def _detect_format_from_filename(filename: str | None) -> LedgerFormat | None:
    """Fallback signal when the header text can't be located: Xero's
    default GL export is named 'General Ledger Summary', MYOB's 'General
    Ledger Report' — so 'summary'/'report' (or an explicit 'xero'/'myob')
    in the filename is a reasonable heuristic."""
    name = (filename or "").lower()
    if "myob" in name:
        return "myob"
    if "xero" in name:
        return "xero"
    if "report" in name:
        return "myob"
    if "summary" in name:
        return "xero"
    return None


def _extract_year(sheet) -> int | None:
    """Best-effort extraction of the financial year's *end* year from a
    period line near the top of the sheet, e.g. Xero's
    'For the period 1 July 2025 to 30 June 2026' -> 2026, or MYOB's
    '01 Apr 2026 - 30 Jun 2026' -> 2026. Takes the SECOND 4-digit year in
    the line, not the first. Returns None if no such line is found —
    callers should treat that as 'ask the user', not silently default to
    something."""
    date_range = re.compile(r"\d{1,2}\s+\w+\s+\d{4}\s*-\s*\d{1,2}\s+\w+\s+(\d{4})")
    for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
        for cell in row:
            if not isinstance(cell, str):
                continue
            if "period" in cell.lower():
                years = re.findall(r"\d{4}", cell)
                if len(years) >= 2:
                    return int(years[1])
                if years:
                    return int(years[0])
            match = date_range.search(cell)
            if match:
                return int(match.group(1))
    return None


def parse_prior_journal(raw_text: str) -> tuple[dict[str, str], int]:
    """Parse pasted, tab-delimited prior-year journal text into a
    Description -> Account lookup.

    Tolerant of blank lines and rows missing a value (skipped, counted).
    Raises ProcessingError if the required columns aren't present at all —
    this is the guard against the wrong text being pasted/passed in.
    """
    cleaned = "\n".join(line for line in raw_text.splitlines() if line.strip())
    if not cleaned:
        raise ProcessingError("Prior journal text is empty.")

    reader = csv.DictReader(io.StringIO(cleaned), delimiter="\t")
    fieldnames = reader.fieldnames or []
    missing = [c for c in PRIOR_JOURNAL_REQUIRED_COLUMNS if c not in fieldnames]
    if missing:
        raise ProcessingError(
            f"This doesn't look like a prior year journal paste — missing "
            f"column(s) {missing}. Columns found: {fieldnames}. Confirm the "
            "right text was pasted before retrying."
        )

    lookup: dict[str, str] = {}
    skipped = 0
    for row in reader:
        description = (row.get("Description") or "").strip()
        account = (row.get("Account") or "").strip()
        if not description or not account:
            skipped += 1
            continue
        lookup[description] = account

    return lookup, skipped
